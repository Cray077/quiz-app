from tkinter import *
from openpyxl import load_workbook
import os

cwd = os.getcwd()  # Get the current working directory (cwd)
files = os.listdir(cwd)  # Get all the files in that directory
print("Files in %r: %s" % (cwd, files))


questionList = []

root = Tk()

root.title = "Quiz App"
root.geometry = "800x600"
root.resizable = False

questionNumberFrame = Frame(root, width=50, height=400)
questionNumberFrame.grid(row=0, column=0, sticky="w", rowspan=2)

questionFrame = Frame(root, pady=20)
questionFrame.grid(row=0, column=1, sticky="n")

answerFrame = Frame(root)
answerFrame.grid(row=1, column=1)

number = IntVar()
questionNumber = Listbox(questionNumberFrame)
questionNumber.grid(row=0, column=0)
questionNumber.bind("<<ListboxSelect>>", number)


        

def display_quiz(question, answer):
    # display the question
    questionLabel = Label(questionFrame, text=question)
    questionLabel.grid(row=0, column=0)
    # display the answer
    for (text, answer) in answer.items():
        radioButton = Radiobutton(answerFrame, text=text, value=answer, width=40, indicatoron=0, padx=10)
        radioButton.pack()

# Load Excel files
wb = load_workbook(cwd+"/quiz.xlsx")
ws = wb.active

run = True
a_column = 1
column = a_column
while run:
    if ws["A"+str(a_column)].value == "start":
        questionList.append(ws["A"+str(column)], ws["B"+str(column)], ws["C"+str(column)], ws["D"+str(column)], ws["E"+str(column)])
    elif ws["A"+str(column)].value == "end":
        run = False
        break
    else: a_column += 1

print(questionList)

question = "Berapakah angka tertinggi dalam 8 bit ?"
answer = {
    "255" : "255",
    "256" : "256",
    "257" : "257",
    "258" : "258",
    "259" : "259"
}

display_quiz(question, answer)

root.mainloop()